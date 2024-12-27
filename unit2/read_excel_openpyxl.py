from openpyxl import load_workbook

workbook = load_workbook("sample3.xlsx")
sheet = workbook.active

# Read data from a specific cell
print(sheet["A1"].value)

# Iterate through rows
for row in sheet.iter_rows(values_only=True): 
    print(row)
